import os
import re
import warnings

import numpy as np
import pandas as pd
from scipy.stats import linregress
import openpyxl
import matplotlib.pyplot as plt
import seaborn as sns
import math
from matplotlib.ticker import ScalarFormatter

warnings.filterwarnings("ignore")

# Constants for flow calculations
VOLUME = 2100
SEATLEAK_VOLUME = 10
COMP_FACTOR = 1
ABS_ZERO = 273.15
STD_PRESSURE = 760

# File paths (modify as needed or override via CLI)
CRAWL_DIR = r"D:\Desktop\Corrosion Test Bench Data\Test 2 June 2025\1st soak"
OUTPUT_DIR = r"D:\Desktop\Corrosion Test Bench Data\Test 2 June 2025\Temperature Sensor Change\DUT Temperature\1st Soak"
ROR_FILENAME = os.path.join(OUTPUT_DIR, "ROR_Results.xlsx")
PRESSURE_FILENAME = os.path.join(OUTPUT_DIR, "Pressure_Results.xlsx")
SEATLEAK_FILENAME = os.path.join(OUTPUT_DIR, "Seatleak_Results.xlsx")



def compute_flow(slope, t_kelvin, volume=VOLUME):
    """Compute flow from slope and temperature (Kelvin), using given volume."""
    return slope * (volume * ABS_ZERO) / (t_kelvin * STD_PRESSURE * COMP_FACTOR)

def ror_global_flow_calc(df: pd.DataFrame):
    """Compute global ROR flow and R² via linear regression."""
    temp_col = next((c for c in df.columns if "Temperature Reading" in c), None)
    if not temp_col:
        raise KeyError("No column found containing 'Temperature Reading'.")

    t_kelvin = df[temp_col].mean() + ABS_ZERO
    slope, _, r_value, _, _ = linregress(df["minute"], df["PT3 Pressure"])
    flow = compute_flow(slope, t_kelvin)
    return flow, r_value

def ror_iterative_flow_calc(df: pd.DataFrame) -> pd.DataFrame:
    """Compute per-row slope and flow using central differences."""
    temp_col = next((c for c in df.columns if "Temperature Reading" in c), None)
    if not temp_col:
        raise KeyError("No column found containing 'Temperature Reading'.")

    t_kelvin = df[temp_col] + ABS_ZERO
    time = df["minute"].values
    pressure = df["PT3 Pressure"].values

    slope = np.gradient(pressure, time, edge_order=2)
    slope[~np.isfinite(slope)] = np.nan

    flow = compute_flow(slope, t_kelvin.values)
    flow[~np.isfinite(flow)] = np.nan

    out = df.copy()
    out["ror_flow_iterative"] = flow
    out["slope_iterative"] = slope
    return out


def seatleak_analysis(df: pd.DataFrame) -> pd.DataFrame:
    """
    Compute seat-leak flow and R for each DUTn on a per-cycle basis,
    using the DUTn Pressure 2 Reading [REAL] and DUTn Temperature Reading [REAL]
    columns, then express flow in sccm and %FS (1000 sccm for DUT1–n-1, 2000 sccm for DUTn).
    Only P2 readings < 2000 are used.
    """
    results = []

    # ensure 'minute' exists
    if "minute" not in df:
        df["minute"] = np.nan

    # find all DUT indices by inspecting the P2 header pattern
    p2_re = re.compile(r"^DUT(\d+)\s+Pressure\s+2\s+Reading\s+\[REAL\]$")
    sensors = sorted(
        int(m.group(1))
        for col in df.columns
        if (m := p2_re.match(col))
    )
    if not sensors:
        # no DUTn Pressure 2 columns found
        return pd.DataFrame(columns=["Cycle","Sensor","Seatleak [sccm]","Seatleak [%FS]","R"])

    # assume the highest DUT# runs at 2000 sccm full scale
    max_dut = sensors[-1]

    for cycle, cycle_df in df.groupby("Cycle", observed=True):
        for sid in sensors:
            p2_col   = f"DUT{sid} Pressure 2 Reading [REAL]"
            temp_col = f"DUT{sid} Temperature Reading [REAL]"

            # skip if either column is missing
            if p2_col not in cycle_df.columns or temp_col not in cycle_df.columns:
                continue

            # filter out any P2 readings ≥ 2000
            filt = cycle_df[cycle_df[p2_col] < 2000]
            if len(filt) < 2:
                # not enough data to fit
                continue

            # linear fit of P2 vs minute
            slope, _, r_value, _, _ = linregress(filt["minute"], filt[p2_col])

            # mean temperature → Kelvin
            t_kelvin = filt[temp_col].mean() + ABS_ZERO

            # compute seat-leak (using the smaller volume)
            seatleak = compute_flow(slope, t_kelvin, volume=SEATLEAK_VOLUME)

            # choose full-scale: 2000 sccm for the last DUT, 1000 sccm otherwise
            fs = 2000.0 if sid == max_dut else 1000.0
            pct_fs = seatleak / fs * 100.0

            results.append({
                "Cycle": cycle,
                "Sensor": sid,
                "Seatleak [sccm]": seatleak,
                "Seatleak [%FS]": pct_fs,
                "R": r_value
            })

    return pd.DataFrame(results)


def pressure_step_analysis(df: pd.DataFrame):
    """
    For each Pressure Setpoint returns:
      1) ref_mean      – mean of PT2 (reference)
      2) dev_means     – mean of each device channel and PT3
      3) pct_error     – % error of each device vs PT2
      4) std_pct_error – std of those % errors across cycles
      5) std_torr      – std of absolute error (torr) across cycles
    """
    assert "Pressure Setpoint" in df.columns, (
        f"'Pressure Setpoint' missing on entry: {df.columns.tolist()}"
    )

    def _trim(g):
        return g.iloc[:-5].tail(30) if len(g) > 5 else pd.DataFrame(columns=g.columns)

    # 1) Trim & group, then filter to START
    df_trimmed = (
        df
        .groupby(["Pressure Setpoint", "Cycle"], group_keys=False)
        .apply(_trim)
        .reset_index(drop=True)
    )
    df_trimmed = df_trimmed[df_trimmed["State"] == "START"]

    # Identify pressure columns
    pressure_cols = [c for c in df_trimmed.columns if "Pressure" in c and c != "Pressure Setpoint"]
    ref_col      = "PT2 Pressure"
    dev_cols     = [c for c in pressure_cols if c != ref_col]  # includes PT3 and DUTs

    # 2) Per-cycle means
    cycle_means = (
        df_trimmed
        .groupby(["Pressure Setpoint", "Cycle"], as_index=False)[pressure_cols]
        .mean()
    )

    # 3) Aggregate means per setpoint
    agg = cycle_means.groupby("Pressure Setpoint", as_index=False)[pressure_cols].mean()
    ref_mean  = agg[["Pressure Setpoint", ref_col]].rename(columns={ref_col: "ref_mean"})
    dev_means = agg[["Pressure Setpoint"] + dev_cols].rename(
        columns={c: f"{c}_mean" for c in dev_cols}
    )

    # 4) Percent error on aggregated means
    pct = agg[["Pressure Setpoint"]].copy()
    for c in dev_cols:
        pct[f"{c}_%err"] = (agg[c] - agg[ref_col]) / agg[ref_col] * 100

    # 5) Std of % error across cycles
    err_cycle = cycle_means.copy()
    for c in dev_cols:
        err_cycle[f"{c}_%err"] = (err_cycle[c] - err_cycle[ref_col]) / err_cycle[ref_col] * 100
    std_pct = (
        err_cycle
        .groupby("Pressure Setpoint", as_index=False)
        .std()[["Pressure Setpoint"] + [f"{c}_%err" for c in dev_cols]]
    ).rename(columns={f"{c}_%err": f"{c}_%err_std" for c in dev_cols})

    # 6) Std of absolute error (torr) across cycles
    err_cycle_torr = cycle_means.copy()
    for c in dev_cols:
        err_cycle_torr[c] = err_cycle_torr[c] - err_cycle_torr[ref_col]
    std_torr = (
        err_cycle_torr
        .groupby("Pressure Setpoint", as_index=False)
        .std()[["Pressure Setpoint"] + dev_cols]
    ).rename(columns={c: f"{c}_std_torr" for c in dev_cols})

    return ref_mean, dev_means, pct, std_pct, std_torr


    
def ror_analysis(df: pd.DataFrame, output_dir: str, file_stem: str):
    """Run ROR analysis, save processed CSV, and return results."""
    low, high    = 30, 350
    pressure_col = "PT3 Pressure"
    dut          = str(int(df["DUT"].mean()))
    test         = df["TEST"].iloc[0]

    # —————————————————————————
    # 1) detect P0, Flow, and Temp channels
    # —————————————————————————
    # try explicit “P0 Pressure Reading [REAL]”
    p0_cols = [c for c in df.columns
               if re.search(r"P0\s+Pressure\s+Reading\s*\[REAL\]$", c)]
    # fallback to generic “Pressure Reading [REAL]” (e.g. DUT16)
    if not p0_cols:
        p0_cols = [c for c in df.columns
                   if re.search(r"Pressure\s+Reading\s*\[REAL\]$", c)]

    flow_cols = [c for c in df.columns
                 if re.search(r"Flow\s+Reading\s*\[REAL\]$", c)]
    temp_cols = [c for c in df.columns
                 if re.search(r"Temperature\s+Reading", c, re.IGNORECASE)]

    if not (p0_cols and flow_cols and temp_cols):
        raise KeyError(
            f"Couldn't detect columns:\n"
            f"  P0   = {p0_cols}\n"
            f"  Flow = {flow_cols}\n"
            f"  Temp = {temp_cols}"
        )

    p0_col        = p0_cols[0]
    mfc_flow_col  = flow_cols[0]
    mfc_temp_col  = temp_cols[0]  # first column containing 'Temperature Reading'
    ror_temp_col  = mfc_temp_col  # use a 'Temperature Reading' column instead of 'PT4 Temperature'

    # —————————————————————————
    # 2) prepare result columns
    # —————————————————————————
    df["ror_flow_global"]    = np.nan
    df["r_squared"]          = np.nan
    df["ror_flow_iterative"] = np.nan
    df["slope_iterative"]    = np.nan

    results   = []
    setpoints = df["ROR Setpoint"].unique()
    cycles    = df["Cycle"].unique()

    # —————————————————————————
    # 3) compute global & iterative flow per setpoint/cycle
    # —————————————————————————
    for cycle in cycles:
        for sp in setpoints:
            mask = (
                (df["ROR Setpoint"] == sp) &
                (df["Cycle"]         == cycle) &
                (df["State"]         == "START") &
                (df[pressure_col]    > low) &
                (df[pressure_col]    < high)
            )
            subset = df.loc[mask]
            if subset.empty:
                continue

            flow, r2 = ror_global_flow_calc(subset)
            df.loc[mask, "ror_flow_global"] = flow
            df.loc[mask, "r_squared"]       = r2

            iter_df = ror_iterative_flow_calc(subset)
            df.loc[iter_df.index, "ror_flow_iterative"] = iter_df["ror_flow_iterative"].values
            df.loc[iter_df.index, "slope_iterative"]    = iter_df["slope_iterative"].values

    # —————————————————————————
    # 4) summarize per setpoint
    # —————————————————————————
    for sp in setpoints:
        mask   = (
            (df["ROR Setpoint"] == sp) &
            (df["State"]         == "START") &
            (df[pressure_col]    > low) &
            (df[pressure_col]    < high)
        )
        subset = df.loc[mask]
        if subset.empty:
            continue

        results.append({
            "time":          subset["Time"].mean().round("D"),
            "dut":           subset["DUT"].mean(),
            "setpoint":      sp,
            "ave_mfc_temp":  subset[mfc_temp_col].mean(),
            "ave_ror_temp":  subset[ror_temp_col].mean(),  # <— was 'PT4 Temperature'
            "ave_p0":        subset[p0_col].mean(),
            "ave_mfc_flow":  subset[mfc_flow_col].mean(),
            "std_mfc_flow":  subset[mfc_flow_col].std(),
            "ave_ror_flow":  subset["ror_flow_global"].mean(),
            "std_ror_flow":  subset["ror_flow_global"].std(),
            "ave_r_squared": subset["r_squared"].mean(),
            "ave_iter_flow": subset["ror_flow_iterative"].mean(),
            "std_iter_flow": subset["ror_flow_iterative"].std(),
        })

    results_df = pd.DataFrame(results)

    # —————————————————————————
    # 5) save processed CSV
    # —————————————————————————
    os.makedirs(output_dir, exist_ok=True)
    processed_csv = os.path.join(output_dir, f"{file_stem}_Processed.csv")
    df.to_csv(processed_csv, index=False)

    return df, results_df, dut, test


def append_and_clean(path: str, sheets: dict):
    def _safe(n: str) -> str:
        # replace forbidden chars and truncate to 31
        n = re.sub(r'[\[\]\*:/\\?]', '_', n)
        return n[:31]

    # 0) load workbook
    wb = openpyxl.load_workbook(path)
    # 1) remove any existing sheets that we’re about to overwrite
    for raw_name in sheets:
        name = _safe(raw_name)
        if name in wb.sheetnames:
            wb.remove(wb[name])
    wb.save(path)

    # 2) append new sheets, replacing if needed
    with pd.ExcelWriter(
        path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace"    # tell pandas to overwrite existing sheets
    ) as writer:
        for raw_name, df in sheets.items():
            writer_name = _safe(raw_name)
            df.to_excel(writer, sheet_name=writer_name, index=False)

    # 3) also clean up any leftover default sheets
    wb = openpyxl.load_workbook(path)
    for orphan in ("Sheet", "Sheet1"):
        if orphan in wb.sheetnames:
            wb.remove(wb[orphan])
    wb.save(path)


def crawl_dir(folder_path: str,
              ror_filename: str,
              pressure_filename: str,
              seatleak_filename: str):
    """Crawl CSVs in a directory, analyze, and write Excel with one sheet per DUT."""
    os.makedirs(os.path.dirname(ror_filename), exist_ok=True)

    # Initialize workbooks
    pd.DataFrame().to_excel(ror_filename, index=False)
    pd.DataFrame().to_excel(pressure_filename, index=False)
    pd.DataFrame().to_excel(seatleak_filename, index=False)

    raw_data = {}
    ror_results = {}
    pressure_results = {}
    seatleak_results = {}
    processed_files = set()

    for file in os.listdir(folder_path):
        if not file.lower().endswith(".csv") or file.lower().endswith("_processed.csv"):
            continue
        fp = os.path.join(folder_path, file)
        if fp in processed_files or not os.path.isfile(fp):
            continue
        processed_files.add(fp)
        file_stem = os.path.splitext(file)[0]

        print(f"Processing file: {file_stem}")
        df = (
            pd.read_csv(fp, low_memory=False)
              .ffill()
              .query("State == 'START'")
              .assign(
                  Time=lambda d: pd.to_datetime(d["Time"], unit="ms", errors="coerce"),
                  minute=lambda d: d["Relative Time (ms)"] / 60000
              )
        )
        
        # ——— Special case: drop Cycle 1 for this specific file ———
        if file_stem == "2025-08-01_22-49_ROR-DUT14-RUN34":
            print(f"[{file_stem}] Callout: excluding Cycle #1 from analysis")
            df = df[df["Cycle"] != 1]
            
            
            
            
        print(f"[{file_stem}] Loaded {len(df)} rows; Test type: {df['TEST'].iloc[0].strip().upper()}")

        raw_data[file_stem] = df.copy()
        avg_time = df["Time"].mean()
        test_type = str(df["TEST"].iloc[0]).strip().upper()

        try:
            if test_type == "ROR":
                print(f"[{file_stem}] ROR analysis")
                _, results_df, _, _ = ror_analysis(df, OUTPUT_DIR, file_stem)
                append_and_clean(ror_filename, {file_stem: results_df})
                ror_results[file_stem] = results_df
                print(f"[{file_stem}] ROR done")

            elif "PRESSURE" in test_type:
                print(f"[{file_stem}] Pressure analysis")
                ref_mean, dev_means, pct_error, std_pct, std_torr = pressure_step_analysis(df)

                # Merge summaries cleanly
                summary = (
                    ref_mean
                    .merge(dev_means, on="Pressure Setpoint")
                    .merge(pct_error, on="Pressure Setpoint")
                    .merge(std_pct, on="Pressure Setpoint")
                    .merge(std_torr, on="Pressure Setpoint")
                )

                # PT3-PT2 diff
                summary["PT3_PT2_diff_mean"] = summary["PT3 Pressure_mean"] - summary["ref_mean"]

                # Absolute error columns (torr)
                for c in dev_means.columns:
                    if c.endswith("_mean"):
                        err = c.replace("_mean", "_err_torr")
                        summary[err] = summary[c] - summary["ref_mean"]

                # Build/write per-DUT sheets
                dut_ids = sorted({int(m.group(1)) for col in dev_means.columns if (m := re.match(r"DUT(\d+)", col))})
                sheets = {}
                for sid in dut_ids:
                    print(f"[{file_stem}] Writing DUT{sid}")
                    c1 = f"DUT{sid} Pressure 1 Reading [REAL]_mean"
                    c2 = f"DUT{sid} Pressure 2 Reading [REAL]_mean"
                    sheet_cols = [
                        "Pressure Setpoint",
                        "ref_mean",
                        "PT3_PT2_diff_mean",
                        c1,
                        c1.replace("_mean", "_err_torr"),
                        c1.replace("_mean", "_%err"),
                        c1.replace("_mean", "_%err_std"),
                        c1.replace("_mean", "_std_torr"),
                        c2,
                        c2.replace("_mean", "_err_torr"),
                        c2.replace("_mean", "_%err"),
                        c2.replace("_mean", "_%err_std"),
                        c2.replace("_mean", "_std_torr"),
                    ]
                    df_dut = summary[sheet_cols].copy()
                    df_dut["Time"] = avg_time
                    sheets[f"DUT{sid}"] = df_dut
                append_and_clean(pressure_filename, sheets)
                pressure_results[file_stem] = sheets
                print(f"[{file_stem}] Pressure done")

            elif "SEATLEAK" in test_type:
                print(f"[{file_stem}] Seatleak analysis")
                seat_df = seatleak_analysis(df)
                summary = (
                    seat_df
                    .groupby("Sensor")
                    .agg({"Seatleak [sccm]": ["mean", "std"], "Seatleak [%FS]": ["mean", "std"], "R": ["mean", "std"]})
                )
                summary.columns = [f"{m}_{s}" for m, s in summary.columns]
                summary = summary.reset_index()
                summary["Time"] = avg_time
                sheets = {str(int(r["Sensor"])): r.to_frame().T for _, r in summary.iterrows()}
                append_and_clean(seatleak_filename, sheets)
                seatleak_results[file_stem] = (seat_df, summary)
                print(f"[{file_stem}] Seatleak done")

            else:
                print(f"[{file_stem}] Unknown TEST: {test_type}")
        except Exception as e:
            print(f"[{file_stem}] Error: {e}")

    print("All files processed.")
    return raw_data, ror_results, pressure_results, seatleak_results







if __name__ == "__main__":
    raw, ror_res, pres_res, seat_res = crawl_dir(
        CRAWL_DIR,
        ROR_FILENAME,
        PRESSURE_FILENAME,
        SEATLEAK_FILENAME
        )

####################### PLOTS #######################


def combine_ror_results(ror_dict):
    """
    ror_dict: { file_stem: results_df, ... }
    where each results_df has arbitrary columns (e.g. ['setpoint','ave_ror_flow', ...]).

    Returns a single DataFrame with:
      - all original columns from each results_df
      - dut       (int)
      - testrun   (int)
    """
    rows = []
    # match both "..._ROR-DUT16-TESTRUN27" and "..._ROR-DUT16-RUN27"
    pattern = re.compile(r".*?_ROR-DUT(\d+)-(?:TESTRUN|RUN)(\d+)$")

    for stem, df in ror_dict.items():
        m = pattern.match(stem)
        if not m:
            # skip any stems that don't fit the pattern
            continue
        dut     = int(m.group(1))
        testrun = int(m.group(2))

        # copy all columns, then add dut & testrun
        tmp = df.copy()
        tmp['dut']     = dut
        tmp['testrun'] = testrun
        rows.append(tmp)

    if not rows:
        # If nothing matched, return an empty DataFrame with the expected columns
        # Attempt to infer column names from any sample df:
        sample_cols = list(next(iter(ror_dict.values())).columns) if ror_dict else []
        return pd.DataFrame(columns=sample_cols + ['dut','testrun'])

    # concatenate all rows into one big DataFrame
    combined = pd.concat(rows, ignore_index=True)
    return combined

def plot_boxplots_per_dut(combined_df, dut, ncols=5, figsize=(15, 4)):
    """
    Open one figure window for the given DUT, with one box-plot per setpoint.

    Parameters
    ----------
    combined_df : pandas.DataFrame
        Must contain columns ['setpoint', 'dut', 'ave_ror_flow'].
    dut : int
        Which DUT number to plot (e.g. 12 or 16).
    ncols : int
        Number of subplots per row.
    figsize : tuple
        Figure size (width, height).
    """
    # filter for this DUT
    df_d = combined_df[combined_df['dut'] == dut]
    setpoints = sorted(df_d['setpoint'].unique())
    n = len(setpoints)

    # determine grid size
    nrows = math.ceil(n / ncols)
    fig, axes = plt.subplots(nrows, ncols, figsize=figsize, sharey=False)
    axes_flat = axes.flatten() if hasattr(axes, "__iter__") else [axes]

    for ax, sp in zip(axes_flat, setpoints):
        df_sp = df_d[df_d['setpoint'] == sp]
        sns.boxplot(
            y='ave_ror_flow',
            data=df_sp,
            ax=ax
        )
        ax.set_title(f"Setpoint {sp}")
        ax.set_xlabel("")  # no x label needed
        ax.set_ylabel("Avg ROR Flow (sccm)")

        # Use a plain formatter and disable the offset (no +1.94e2)
        formatter = ScalarFormatter(useOffset=False)
        formatter.set_scientific(False)
        ax.yaxis.set_major_formatter(formatter)

    # remove any unused subplots
    for ax in axes_flat[n:]:
        fig.delaxes(ax)

    fig.suptitle(f"DUT {dut} — ROR Flow by Setpoint, Ambient 7-Day", fontsize=16)
    plt.tight_layout(rect=[0, 0, 1, 0.95])
    plt.show()

def plot_mfc_flow(df):
    """
    Plot MFC flow vs setpoint, colored and styled by DUT.
    
    Parameters
    ----------
    df : pandas.DataFrame
        Must contain columns:
          - 'time'           (str or datetime)
          - 'dut'            (int or str)
          - 'setpoint'       (numeric)
          - 'ave_mfc_flow'   (numeric)
    """
    # Define some linestyles/markers to cycle through
    linestyles = ['-', '--', ':', '-.']
    markers    = ['o', 's', '^', 'd', 'v', 'x', '+']
    
    fig, ax = plt.subplots(figsize=(8,5))
    
    # Loop over each DUT
    for idx, dut in enumerate(sorted(df['dut'].unique())):
        sub = df[df['dut'] == dut]
        ls = linestyles[idx % len(linestyles)]
        mk = markers[idx % len(markers)]
        ax.plot(sub['setpoint'], sub['ave_mfc_flow'],
                label=f'DUT {dut}',
                linestyle=ls,
                marker=mk)
    
    # Build title from the unique time(s)
    times = df['time'].unique()
    if len(times) == 1:
        time_str = times[0]
    else:
        time_str = ', '.join(map(str, times))
    
    ax.set_xlabel('Setpoint')
    ax.set_ylabel('Average MFC Flow')
    ax.set_title(f'MFC Flow vs Setpoint ({time_str})')
    ax.legend(title='DUT', loc='best')
    ax.grid(True)
    fig.tight_layout()
    plt.show()
    
def plot_std_ror_flow(df):
    """
    Scatter-plot of ROR flow standard deviation vs setpoint, one marker per DUT.

    Parameters
    ----------
    df : pandas.DataFrame
        Must contain columns:
          - 'time'           (str or datetime)
          - 'dut'            (int or str)
          - 'setpoint'       (numeric)
          - 'std_ror_flow'   (numeric)
    """
    markers = ['o', 's', '^', 'd', 'v', 'x', '+']
    
    # Font-size settings
    title_fs   = 16
    label_fs   = 14
    tick_fs    = 12
    legend_fs  = 12
    legend_tfs = 13  # legend title font size
    
    fig, ax = plt.subplots(figsize=(8,5))
    
    for idx, dut in enumerate(sorted(df['dut'].unique())):
        sub = df[df['dut'] == dut]
        mk = markers[idx % len(markers)]
        ax.scatter(
            sub['setpoint'],
            sub['std_ror_flow'],
            label=f'DUT {dut}',
            marker=mk,
            s=60  # marker size
        )
    
    # Build title from the unique time(s)
    times = df['time'].unique()
    if len(times) == 1:
        time_str = times[0]
    else:
        time_str = ', '.join(map(str, times))
    
    ax.set_xlabel('Setpoint', fontsize=label_fs)
    ax.set_ylabel('1 Sigma Flow Repeatability [sccm]', fontsize=label_fs)
    ax.set_title(f'ROR Flow Repeatability vs Setpoint, 2025.06.24 Post Passivation Retake', fontsize=title_fs)
    
    ax.tick_params(axis='both', which='major', labelsize=tick_fs)
    
    legend = ax.legend(title='DUT', loc='best', fontsize=legend_fs)
    legend.get_title().set_fontsize(legend_tfs)
    
    ax.grid(True)
    fig.tight_layout()
    plt.show()

combined = combine_ror_results(ror_res)
# plot_mfc_flow(combined)
plot_std_ror_flow(combined)
# # Usage for DUT12 and DUT16:
# plot_boxplots_per_dut(combined, dut=12, ncols=5, figsize=(15, 3))
# plot_boxplots_per_dut(combined, dut=16, ncols=5, figsize=(15, 3))


def check_r_squared(df, threshold=0.99):
    """
    Checks the 'ave_r_squared' column of df and prints a prominent warning
    for any values below the given threshold.
    
    Parameters
    ----------
    df : pandas.DataFrame
        Must contain a column named 'ave_r_squared'. May also contain
        'dut', 'testrun', and 'setpoint' for context in the warning.
    threshold : float, optional
        The R² cutoff below which to warn. Default is 0.99.
    """
    # Find all rows where ave_r_squared is below threshold
    low_r2 = df[df['ave_r_squared'] < threshold]
    
    if not low_r2.empty:
        banner = "*" * 80
        print("\n" + banner)
        print(f"!!!  WARNING: {len(low_r2)} R² VALUE(S) BELOW {threshold:.2f} DETECTED  !!!")
        print(banner)
        # Print details for each failing row
        for idx, row in low_r2.iterrows():
            context = []
            for col in ('dut', 'testrun', 'setpoint'):
                if col in df.columns:
                    context.append(f"{col}={row[col]}")
            context_str = ", ".join(context) if context else f"row index {idx}"
            print(f"  • {context_str} → ave_r_squared = {row['ave_r_squared']:.4f}")
        print(banner + "\n")
    else:
        print(f"All ave_r_squared values are ≥ {threshold:.2f} ✅")

check_r_squared(combined)

def combine_seatleak_summaries(seatleak_dict):
    """
    Given seatleak_dict: { file_stem: (full_df, summary_df), … },
    returns one DataFrame with all the summary_dfs stacked together,
    and a column 'file_stem' so you know which run each row came from.
    """
    df_list = []
    for file_stem, (_, summary_df) in seatleak_dict.items():
        tmp = summary_df.copy()
        tmp['file_stem'] = file_stem
        df_list.append(tmp)
    combined = pd.concat(df_list, ignore_index=True)
    return combined

# combine only the summary dfs from seat_res
combined_seat_summary = combine_seatleak_summaries(seat_res)


def plot_seatleak_boxplot_all_sensors(combined_summary, figsize=(12, 6)):
    """
    Boxplot of Seatleak [%FS]_mean for each sensor, aggregated across all runs.
    
    Parameters
    ----------
    combined_summary : pandas.DataFrame
        Must contain columns ['Sensor_', 'Seatleak [%FS]_mean'].
    figsize : tuple
        Size of the figure, e.g. (width, height).
    """
    plt.figure(figsize=figsize)
    sns.boxplot(
        data=combined_summary,
        x='Sensor_',
        y='Seatleak [%FS]_mean'
    )
    plt.xlabel('Sensor')
    plt.ylabel('Seatleak [% of Full Scale] (mean)')
    plt.title('Distribution of Seatleak [%FS] Mean Across All Runs by Sensor')
    plt.tight_layout()
    plt.show()
    
# plot_seatleak_boxplot_all_sensors(combined_seat_summary)


def plot_pressurestep_error(pres_res, filename, sub_idx=2):
    """
    pres_res : dict
        { filename: (df0, df1, df2), … }
    filename : str
        Top‐level key in pres_res
    sub_idx : int
        Always 2 for the error summary DataFrame
    """
    # Grab the error summary sheet
    try:
        df_err = pres_res[filename][sub_idx]
    except KeyError:
        raise KeyError(f"File {filename!r} not in pres_res; available: {list(pres_res.keys())}")
    except IndexError:
        raise IndexError(f"pres_res[{filename}] has no element at index {sub_idx}")

    # Unique setpoints
    sps = sorted(df_err['Pressure Setpoint'].unique())

    # All the "Pressure 1" %err columns
    cols1 = [c for c in df_err.columns
             if 'Pressure 1' in c and c.endswith('_%err')]

    # All the "Pressure 2" %err columns
    cols2 = [c for c in df_err.columns
             if 'Pressure 2' in c and c.endswith('_%err')]

    def _plot_box(cols, title):
        # collect one list per setpoint
        data = [
            df_err.loc[df_err['Pressure Setpoint']==sp, cols].values.flatten()
            for sp in sps
        ]
        fig, ax = plt.subplots()
        ax.boxplot(data, labels=sps, showfliers=False)
        ax.set_xlabel('Pressure Setpoint')
        ax.set_ylabel('Error [%RDG]')
        ax.set_title(f"{title}\n All DUTs")
        fig.tight_layout()

    _plot_box(cols1, 'Pressure 1 % Error by Setpoint')
    _plot_box(cols2, 'Pressure 2 % Error by Setpoint')
    
    plt.show()

# plot_pressurestep_error(
#     pres_res,
#     '2025-06-04_17-54_PRESSURESTEP-DUT16-RUN26')

